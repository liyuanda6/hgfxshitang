# -*- coding: utf-8 -*-
"""生成《附小学生餐费核对表》多工作表 xlsx（每班一页）。
数据源：
  1) 工作安排表 xlsx -> 学校开设的全部班级（转成「一年级1班」形式）
  2) 就餐系统 /api/state -> 学生、用餐类型、天数、餐费、备注
版式：第一行合并标题；列为 序号/学生姓名/用餐类型/每日餐费/用餐天数/总费用/备注；
      总费用=当月真实费用；底部班主任审核、附小后勤负责人审核、日期留空。
"""
import json
import re
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

XLSX_SRC = r'C:\Users\lyd\Downloads\（终）2026秋附小工作安排表简版 1.xlsx'
STATE_JSON = r'C:\Users\lyd\WorkBuddy\2026-09-05-15-25-09\class-meal-system\state-tmp.json'
OUT = r'C:\Users\lyd\Downloads\附小学生餐费2026年9月.xlsx'

MONTH_LABEL = '2026年9月'
SCHOOL = '黄冈中学附属小学'

CN = {'一': 1, '二': 2, '三': 3, '四': 4, '五': 5, '六': 6,
      '七': 7, '八': 8, '九': 9, '十': 10}

def cn2int(s):
    s = s.strip()
    if s in CN:
        return CN[s]
    if s.startswith('十'):
        return 10 + (CN.get(s[1:], 0) if len(s) > 1 else 0)
    if '十' in s:
        a, b = s.split('十')
        return (CN.get(a, 1) if a else 1) * 10 + (CN.get(b, 0) if b else 0)
    return int(s)

def parse_meal_class(name):
    """一年级一班 -> (1,1)"""
    m = re.match(r'^([一二三四五六七八九十]+)年级([一二三四五六七八九十]+)班$', name)
    if m:
        return cn2int(m.group(1)), cn2int(m.group(2))
    return None

def parse_xlsx_class(name):
    """一（1）/ 三（5）班 -> (1,1)/(3,5)"""
    m = re.match(r'^([一二三四五六七八九十]+)[（(](\d+)[）)]班?$', name)
    if m:
        return cn2int(m.group(1)), int(m.group(2))
    # 兼容「一年级1班」直接写法
    m2 = re.match(r'^([一二三四五六七八九十]+)年级(\d+)班$', name)
    if m2:
        return cn2int(m.group(1)), int(m.group(2))
    return None

def canonical(grade, cls):
    """(1,1) -> 一年级1班"""
    g = [k for k, v in CN.items() if v == grade][0]
    return f'{g}年级{cls}班'

def title_text(grade, cls):
    g = [k for k, v in CN.items() if v == grade][0]
    return f'{MONTH_LABEL}{SCHOOL} {g} 年级  {cls}  班学生餐费核对表'

# ---- 1. 读取工作安排表，拿到全部班级（grade, cls）----
wb_src = openpyxl.load_workbook(XLSX_SRC, data_only=True)
all_classes = []  # list of (grade, cls)
seen = set()
for ws in wb_src.worksheets:
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell and isinstance(cell, str):
                p = parse_xlsx_class(cell.strip())
                if p and p not in seen:
                    seen.add(p)
                    all_classes.append(p)
all_classes.sort()
print('工作安排表中的班级数:', len(all_classes))
for g, c in all_classes:
    print('  ', canonical(g, c))

# ---- 2. 读取就餐系统状态 ----
state = json.load(open(STATE_JSON, encoding='utf-8'))
settings = state.get('settings', {})
bf = settings.get('breakfastPrice', 0)
lf = settings.get('lunchPrice', 0)
std_label = {s['key']: s['label'] for s in state.get('standards', [])}
default_std = state.get('defaultStandard', 'BL')
days_map = state.get('days', {})
records = state.get('records', {}).get('2026-09', {})
students = state.get('students', [])
classes = state.get('classes', [])

def daily_fee(std):
    f = 0
    if 'B' in std:
        f += bf
    if 'L' in std:
        f += lf
    return f

# 建立 班级键 -> 学生列表
cls_students = {}
for s in students:
    cls_students.setdefault(s['classId'], []).append(s)

# 建立 就餐班级 (grade,cls) -> 班级对象
meal_by_key = {}
for c in classes:
    k = parse_meal_class(c['name'])
    if k:
        meal_by_key[k] = c

# ---- 3. 样式 ----
thin = Side(style='thin', color='000000')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
title_font = Font(name='宋体', bold=True, size=15)
hdr_font = Font(name='宋体', bold=True, size=11)
cell_font = Font(name='宋体', size=11)
bold_font = Font(name='宋体', bold=True, size=11)
hdr_fill = PatternFill('solid', fgColor='DDEBF7')
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left = Alignment(horizontal='left', vertical='center')
right = Alignment(horizontal='right', vertical='center')
money_fmt = '"¥"#,##0.00'

wb = openpyxl.Workbook()
wb.remove(wb.active)

COLS = 7  # A..G
col_w = [6, 12, 12, 11, 10, 13, 26]

def build_sheet(ws, grade, cls, student_list, class_obj):
    # 列宽
    for i in range(COLS):
        ws.column_dimensions[get_column_letter(i + 1)].width = col_w[i]
    # 标题
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=COLS)
    t = ws.cell(row=1, column=1, value=title_text(grade, cls))
    t.font = title_font
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    # 表头
    headers = ['序号', '学生姓名', '用餐类型', '每日餐费', '用餐天数', '总费用', '备注']
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=j, value=h)
        c.font = hdr_font
        c.alignment = center
        c.fill = hdr_fill
        c.border = border
    ws.row_dimensions[2].height = 22

    if student_list:
        days = days_map.get(f"{class_obj['id']}|2026-09", 0)
        r = 3
        for idx, s in enumerate(student_list, start=1):
            rec = records.get(s['id'], {})
            std = rec.get('standard') or default_std
            deduction = rec.get('deduction', 0) or 0
            remark = (rec.get('remark') or '').strip()
            df = daily_fee(std)
            gross = days * df
            real = max(gross - deduction, 0)
            vals = [idx, s['name'], std_label.get(std, std), df, days, real, remark]
            for j, v in enumerate(vals, start=1):
                c = ws.cell(row=r, column=j, value=v)
                c.font = cell_font
                c.border = border
                if j == 1:
                    c.alignment = center
                elif j in (2, 7):
                    c.alignment = left
                elif j == 3 or j == 5:
                    c.alignment = center
                else:  # 4,6 费用
                    c.alignment = right
                    c.number_format = money_fmt
            r += 1
        last = r - 1
        # 合计行
        ws.cell(row=r, column=1, value='合计').font = bold_font
        ws.cell(row=r, column=1).alignment = center
        ws.cell(row=r, column=1).border = border
        # 中间列留空带边框
        for j in range(2, 6):
            cc = ws.cell(row=r, column=j, value='')
            cc.border = border
            cc.alignment = center
        tot = ws.cell(row=r, column=6, value=f'=SUM(F3:F{last})')
        tot.font = bold_font
        tot.alignment = right
        tot.number_format = money_fmt
        tot.border = border
        ws.cell(row=r, column=7, value='').border = border
        ws.row_dimensions[r].height = 20
        data_end = r
    else:
        # 空模板：提示一行
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=COLS)
        nt = ws.cell(row=3, column=1, value='（本班暂无就餐登记数据，可手动填写）')
        nt.font = Font(name='宋体', size=10, italic=True, color='888888')
        nt.alignment = center
        for j in range(1, COLS + 1):
            ws.cell(row=3, column=j).border = border
        # 合计 0
        r = 4
        ws.cell(row=r, column=1, value='合计').font = bold_font
        ws.cell(row=r, column=1).alignment = center
        ws.cell(row=r, column=1).border = border
        for j in range(2, 7):
            ws.cell(row=r, column=j, value='').border = border
        tot = ws.cell(row=r, column=6, value=0)
        tot.font = bold_font
        tot.alignment = right
        tot.number_format = money_fmt
        tot.border = border
        ws.cell(row=r, column=7, value='').border = border
        data_end = r

    # 底部签名行（留两行空白间距）
    sig1 = data_end + 2
    ws.merge_cells(start_row=sig1, start_column=1, end_row=sig1, end_column=3)
    a = ws.cell(row=sig1, column=1, value='班主任审核：________________')
    a.font = cell_font
    a.alignment = left
    sig2 = sig1 + 1
    ws.merge_cells(start_row=sig2, start_column=1, end_row=sig2, end_column=3)
    b = ws.cell(row=sig2, column=1, value='附小后勤负责人审核：________________')
    b.font = cell_font
    b.alignment = left
    # 日期靠右留空
    ws.merge_cells(start_row=sig2, start_column=5, end_row=sig2, end_column=7)
    d = ws.cell(row=sig2, column=5, value='日期：        年      月      日')
    d.font = cell_font
    d.alignment = right

    # 冻结表头
    ws.freeze_panes = 'A3'

for (grade, cls) in all_classes:
    name = canonical(grade, cls)
    meal = meal_by_key.get((grade, cls))
    ws = wb.create_sheet(title=name[:31])
    if meal:
        slist = cls_students.get(meal['id'], [])
        build_sheet(ws, grade, cls, slist, meal)
    else:
        build_sheet(ws, grade, cls, [], None)

wb.save(OUT)
print('\n已生成:', OUT)
print('工作表数量:', len(wb.sheetnames))
